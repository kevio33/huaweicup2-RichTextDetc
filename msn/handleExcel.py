import re
from openpyxl import load_workbook
import os

#读excel
def readExcel(file_path):
    wb = load_workbook(file_path)

    mappinglist = []
    for x in wb.sheetnames:
        sheet = wb[x]
        column_names = []
        for cell in sheet[1]:
            column_names.append(cell.value)
        columnnamecell=sheet['1'] #获取第一行的所有cell
        columnname=[y.value for y in columnnamecell ]   #字段在excel中的行数列表
        for rownum in range(2,sheet.max_row+1):
            mapping = {}
            for x in column_names:
                index_num=columnname.index(x)+1  #excel索引和数组索引差1
                mapping[x]=sheet.cell(rownum,index_num).value
            mappinglist.append(mapping)

    wb.close()

    return mappinglist

def handleExcel(file_name,file_path):
    #msn原代码
    # sensitive_columns = ["用户名", "设备地址", "密码"]
    # wb = load_workbook(filePath)
    # sheet = wb.active
    # # 获取列名和对应的索引
    # col_indices = {}
    # for col_num, col_cells in enumerate(sheet.iter_cols(values_only=True)):
    #     if col_cells[0] in sensitive_columns:
    #         col_indices[col_cells[0]] = col_num
    # sensitive_data = []

    # for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第2行开始，因为第1行是列名
    #     for col_name, col_num in col_indices.items():
    #         if row[col_num] is not None:
    #             sensitive_data.append(row[col_num])

    # current_directory = os.path.dirname(os.path.abspath(__file__))
    # output_path = os.path.join(current_directory, "outputExcel.txt")
    # save_to_txt(sensitive_data, output_path)

    #--------------kevin修改的代码----------------
    lis = readExcel(file_path)
    current_directory = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(current_directory, "{}.txt".format(file_name))
    save_to_txt(str(lis), output_path, file_name)


#保存结果
def save_to_txt(data, output_path, file_name):
    dict = {}
    with open(output_path, 'w') as file:
        # for item in data:
        #     file.write(item + '\n')
        dict[file_name] = data
        file.write(str(dict) + '\n')

# if __name__ == '__main__':
#     handleExcel('资产梳理.xlsx','D:\huaweicup\huaweicup2-RichTextDetc\赛题材料\office\资产梳理.xlsx')
